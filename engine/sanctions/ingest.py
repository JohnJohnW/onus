"""Parse a sanctions list (DFAT Consolidated List) and import it as a versioned
snapshot. The parser is header-tolerant: columns are matched by header substring
and the full raw row is retained, so it survives DFAT reordering or renaming
columns and works for both CSV and XLSX. Names for one Reference may span several
rows (DFAT lists aliases as extra rows); those are grouped into one entry."""
from __future__ import annotations

import csv
import hashlib
import io
import json
from datetime import datetime, timezone
from typing import Optional

from sqlalchemy import update
from sqlalchemy.orm import Session

from models import SanctionsEntry, SanctionsListVersion
from sanctions.matching import normalize_name


def _pick(row: dict, *needles: str) -> str:
    for needle in needles:
        for key, value in row.items():
            if needle in key:
                return (value or "").strip()
    return ""


def rows_to_entries(rows: list[dict]) -> list[dict]:
    """Group raw header->value rows into entry dicts keyed by Reference."""
    grouped: dict[str, dict] = {}
    order: list[str] = []
    for raw in rows:
        norm = {(k or "").strip().lower(): ("" if v is None else str(v)) for k, v in raw.items()}
        reference = _pick(norm, "reference", "id")
        name = _pick(norm, "name of individual", "name")
        if not name and not reference:
            continue
        key = reference or name
        if key not in grouped:
            grouped[key] = {
                "reference": reference,
                "entity_type": (_pick(norm, "type") or "unknown").lower(),
                "primary_name": name,
                "_names": [],
                "dob": _pick(norm, "date of birth", "dob", "birth"),
                "place_of_birth": _pick(norm, "place of birth"),
                "citizenship": _pick(norm, "citizenship", "nationality"),
                "address": _pick(norm, "address"),
                "listing_info": _pick(norm, "listing", "committee", "additional"),
                "raw": raw,
            }
            order.append(key)
        if name:
            grouped[key]["_names"].append(name)

    entries: list[dict] = []
    for key in order:
        entry = grouped[key]
        names = entry.pop("_names")
        if not entry["primary_name"] and names:
            entry["primary_name"] = names[0]
        search, seen = [], set()
        for raw_name in [entry["primary_name"], *names]:
            normalized = normalize_name(raw_name)
            if normalized and normalized not in seen:
                seen.add(normalized)
                search.append(normalized)
        entry["search_names"] = search
        entry["aliases"] = [n for n in names if n and n != entry["primary_name"]]
        entries.append(entry)
    return entries


def parse_csv(text: str) -> list[dict]:
    return list(csv.DictReader(io.StringIO(text)))


def parse_xlsx(content: bytes) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - only hit if the dep is missing
        raise RuntimeError("openpyxl is required to parse XLSX sanctions lists.") from exc
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    header_row = next(rows_iter, None) or []
    headers = [str(h).strip() if h is not None else f"col{i}" for i, h in enumerate(header_row)]
    out: list[dict] = []
    for row in rows_iter:
        if row is None or all(c is None for c in row):
            continue
        out.append({headers[i]: row[i] for i in range(min(len(headers), len(row)))})
    return out


def content_hash(entries: list[dict]) -> str:
    payload = json.dumps(
        [[e.get("reference"), e.get("primary_name"), e.get("search_names")] for e in entries],
        sort_keys=True,
        ensure_ascii=False,
    )
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def import_version(
    db: Session, *, source: str, origin: str, entries: list[dict], note: Optional[str] = None
) -> SanctionsListVersion:
    """Persist entries as a new current snapshot, demoting prior versions of the
    same source. Returns the new version row."""
    version = SanctionsListVersion(
        source=source,
        origin=origin,
        fetched_at=datetime.now(timezone.utc),
        content_hash=content_hash(entries),
        entry_count=len(entries),
        is_current=True,
        note=note,
    )
    db.execute(
        update(SanctionsListVersion)
        .where(SanctionsListVersion.source == source, SanctionsListVersion.is_current.is_(True))
        .values(is_current=False)
    )
    db.add(version)
    db.flush()
    for entry in entries:
        db.add(
            SanctionsEntry(
                version_id=version.id,
                reference=entry.get("reference") or None,
                entity_type=entry.get("entity_type") or "unknown",
                primary_name=entry.get("primary_name") or "",
                search_names=entry.get("search_names") or [],
                aliases=entry.get("aliases") or [],
                dob=entry.get("dob") or None,
                place_of_birth=entry.get("place_of_birth") or None,
                citizenship=entry.get("citizenship") or None,
                address=entry.get("address") or None,
                listing_info=entry.get("listing_info") or None,
                raw=entry.get("raw"),
            )
        )
    db.commit()
    db.refresh(version)
    return version
